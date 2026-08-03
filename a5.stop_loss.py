"""
a4-style driver: purchase-anchored stop-loss / upside / z monitor for the
portfolio held in your worksheet.

Roadmap parity with a4.credit_rating_daily:
    1. gather tickers + anchor        (a4: latest CSV per ticker  ->  here: latest
                                        yyyy.mm.dd worksheet -> avg cost per position)
    2. fetch OHLC per ticker via FMP  (same FMPClient.get_data call as a4)
    3. compute key metrics            (a4: zscore_stats decline triggers -> here:
                                        position_stops, anchored on each Avg cost)
    4. save CSVs                      (a4: _triggered_{date}.csv -> here: a full
                                        summary plus a _triggered_ file)
"""

import os
import re
import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# --- Absolute path fix (mirrors a4) so _src modules import under CI ---
current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.join(current_dir, "_src")
sys.path.insert(0, src_dir)

from financial_tools import FMPClient
from zscore_stats import ZScoreConfig, prepare_price_df
from position_stops import augment_positions, summarize_positions

pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
pd.set_option("display.expand_frame_repr", False)

# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------
WORKSHEET_PATH = "worksheet.xlsx"
OUTPUT_DIR = Path("bqr")
EXCLUDE_TICKERS = ["WOLF", "IBKR"]      # non-marketable / fractional-only names

STOP_PCT = 0.20                                     # stop-loss @ 20% below avg cost
UP_PCT = 0.20                                       # "return high" @ 20% above avg cost
Z_UP_THRESHOLDS = (1.0, 1.5, 2.0, 2.5)              # upward z levels
SAVE_DAILY = False                                  # also dump the full daily detail
GIT_IGNORE = True                                   # skip the output dir in git

CONFIG = ZScoreConfig(
    volatility_returns=252,
    horizon=20,
    thresholds=(1.0, 1.5, 2.0, 2.5),               # base module needs non-empty; upside is added in position_stops
    decline_thresholds=(0.10, 0.15, 0.20),
)

SHEET_RE = re.compile(r"^\d{4}\.\d{2}\.\d{2}$")

# a4-style short column names for the saved summary
SUMMARY_RENAME = {
    "symbol": "symbol",
    "purchase_date": "purchase date",
    "avg_cost": "avg cost",
    "sigma (%)": "sigma (%)",
    'last_date': "Last Date",
    "cum_return_from_purchase (%)": "Cum ret (%)",
    "peak_high_return (%)": "Peak high (%)",

    "stop_price": "down 20pct px",
    "stop_hit": "down 20pct",
    "stop_first_date": "down 20pct dt",

    "high_target": "up 20pct px",
    "high_hit": "up 20pct",
    "high_first_date": "up 20pct dt",

    "trail_stop_trigger": "trail stop",
    "trail_stop_first_date": "trail stop dt",

    "z_up_trigger_1_0": "z1.0 up",
    "z_up_1_0_first_date": "z1.0 up dt",
    "z_up_trigger_1_5": "z1.5 up",
    "z_up_1_5_first_date": "z1.5 up dt",
    "z_up_trigger_2_0": "z2.0 up",
    "z_up_2_0_first_date": "z2.0 up dt",
    "z_up_trigger_2_5": "z2.5 up",
    "z_up_2_5_first_date": "z2.5 up dt",
}

# ------------------------------------------------------------------
# 1. Load positions from the latest yyyy.mm.dd worksheet -> avg frame
#    (your worksheet-reading + weighted-avg-cost code, wrapped)
# ------------------------------------------------------------------
def load_positions(path: str) -> tuple[pd.DataFrame, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    name = max(n for n in wb.sheetnames if SHEET_RE.match(n))     # newest date sheet
    rows = wb[name].iter_rows(min_row=2, min_col=2, values_only=True)  # start at B2

    header = next(rows)
    headers = []
    for v in header:                       # walk right until an empty header
        if v is None or str(v).strip() == "":
            break
        headers.append(str(v).strip())

    data = []
    for row in rows:                       # walk down until a blank row
        cells = row[:len(headers)]
        if all(c is None or str(c).strip() == "" for c in cells):
            break
        data.append(cells)
    wb.close()

    df = pd.DataFrame(data, columns=headers)

    def parse_cb(s):
        m = re.search(r"\$([\d,]+\.?\d*)\s*/\s*([\d.]+)\s*sh", str(s))
        return pd.Series({"Cost basis": float(m.group(1).replace(",", "")),
                          "Quantity":   float(m.group(2))})

    parsed = df["Cost basis + quantity"].apply(parse_cb)
    out = pd.concat([df[["Ticker", "Purchase date"]], parsed], axis=1)

    avg = (out.groupby(["Ticker", "Purchase date"], as_index=False)
              .agg({"Cost basis": "sum", "Quantity": "sum"}))
    avg["Avg cost"] = (avg["Cost basis"] / avg["Quantity"]).round(4)
    avg = avg[~avg["Ticker"].isin(EXCLUDE_TICKERS)]
    avg["Purchase date"] = pd.to_datetime(avg["Purchase date"]).dt.strftime("%Y-%m-%d")
    avg = avg.sort_values("Purchase date", ascending=False).reset_index(drop=True)
    return avg, name


# ------------------------------------------------------------------
# 2. Fetch OHLC per ticker -> combined price_df (module format)
#    Same FMP call as a4; validate per-symbol so one bad name can't sink the run.
# ------------------------------------------------------------------
def build_price_df(tickers, client) -> tuple[pd.DataFrame, list[dict]]:
    frames, errors = [], []
    for sym in dict.fromkeys(tickers):          # unique tickers, order-preserving (multi-lot safe)
        raw = client.get_data("historical-price-eod", sym)
        if not raw:
            errors.append({"symbol": sym, "error": "no data returned from FMP"})
            continue
        try:
            frame = pd.DataFrame(raw)
            frame["symbol"] = sym                      # guarantee the symbol column
            frames.append(prepare_price_df(frame))     # per-symbol validation/isolation
        except Exception as exc:
            errors.append({"symbol": sym, "error": f"price validation: {exc}"})
    price_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return price_df, errors


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def main(worksheet_path: str = WORKSHEET_PATH,
         client=None,
         config: ZScoreConfig = CONFIG,
         output_dir: Path = OUTPUT_DIR) -> pd.DataFrame:
    output_dir.mkdir(parents=True, exist_ok=True)
    client = client or FMPClient()

    # 1 ----------------------------------------------------------------
    avg, sheet_name = load_positions(worksheet_path)
    print(f"Loaded {len(avg)} positions from sheet {sheet_name}")
    print(avg.to_string(index=False))

    # 2 ----------------------------------------------------------------
    price_df, fetch_errors = build_price_df(avg["Ticker"].tolist(), client)

    # 3 ----------------------------------------------------------------
    daily, calc_errors = augment_positions(
        price_df, avg, config,
        stop_pct=STOP_PCT, up_pct=UP_PCT, z_up_thresholds=Z_UP_THRESHOLDS,
    )

    errors = pd.DataFrame(fetch_errors + calc_errors.to_dict("records"))
    if not errors.empty:
        print("\nSkipped tickers:")
        print(errors.to_string(index=False))

    summary = summarize_positions(daily, z_up_thresholds=Z_UP_THRESHOLDS)
    if summary.empty:
        print("\nNo positions produced statistics.")
        return summary

    # 4 ----------------------------------------------------------------
    display = summary.rename(columns=SUMMARY_RENAME).sort_values(["purchase date"])

    if GIT_IGNORE:
        display = display.drop(columns=[
            "Last Date", "Cum ret (%)", "Peak high (%)", "down 20pct px",  "up 20pct px"])

    summary_file = output_dir / f"_exit_triggered_.csv"
    display.to_csv(summary_file, index=False)
    print(f"\nSaved {summary_file}")

    if SAVE_DAILY and not daily.empty:
        daily_file = output_dir / f"_holding_daily_.csv"
        daily.to_csv(daily_file, index=False)
        print(f"Saved {daily_file}")

    return summary

if __name__ == "__main__":
    main()